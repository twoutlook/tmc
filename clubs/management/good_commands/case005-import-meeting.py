from django.core.management.base import BaseCommand, CommandError
from django.db.models import Count,Sum
# from mach.models import MachTask
from case005.models import Data2

from case005.models import Club
from case005.models import Person
from case005.models import Role
from case005.models import Meeting


import os
import array
import openpyxl
import statistics as stats

from openpyxl import load_workbook # to read

     # https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-one-cell
from openpyxl import Workbook # to write

# 2019-6-15
# 為什麼發現 扣除未流轉，沒能更新到表單？
class Command(BaseCommand):
   
        
    def handle(self, *args, **options):
     
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        self.stdout.write(BASE_DIR)

        src = os.path.join(BASE_DIR, 'SAA-201906-201911.xlsx')
    
      
        self.stdout.write(src)
        
        self.stdout.write(self.style.SUCCESS('This program is "%s"' % 'case002-import-guest.py'))
        
        # https://openpyxl.readthedocs.io/en/stable/optimized.html
        
        wb = load_workbook(filename=src,read_only=False)
        # ws = wb['6.4']
        
        


      

        def m3(src,sheetname):
            wb = load_workbook(filename=src,read_only=False)
        
            ws =wb[sheetname]
            
            def get_val(cell):
                if cell == 'None':
                    return '---'
                if row[1].value in [None,'None']:
                    return 'xxx'
                else:
                    if str(cell.value) == 'None':
                        return '---'
                    
                    return str(cell.value)
            rec = 0
            mem_cnt = 0
            for row in ws.rows:
                rec += 1
                    
                A = get_val(row[0])
                # self.stdout.write(A)
                B = get_val(row[1])
                #  4 if var1 is None else var1
                C = get_val(row[2])
                
                D = get_val(row[3])
                E = get_val(row[4])
                F = get_val(row[5])
                G = get_val(row[6])
                H = get_val(row[7])
                I = get_val(row[8])
                J = get_val(row[9])
                K = get_val(row[10])
               
                if B == 'Member':
                    mem_cnt += 1
              
                    # self.stdout.write('****')
                    # self.stdout.write(str(rec))
                    # self.stdout.write('****')
                    # self.stdout.write(A)
                    # self.stdout.write(B)
                    # self.stdout.write(C)
                    # self.stdout.write(D)
                    # self.stdout.write(E)
                    # self.stdout.write(F)
                    # self.stdout.write(G)
                    # self.stdout.write(H)
                    # self.stdout.write(I)
                    # self.stdout.write(J)
                    # self.stdout.write(K)
                
                    temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C+' '+ E+' '+ G
                    # temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C
                    self.stdout.write(temp)
                    # obj = Data2(name=A,member='Member',date1='2019-06-05',role=C)
                    # obj.save()


                    def addData2(name,member, date1, role):
                        obj = Data2(name=name,member=member,date1=date1,role=role)
                        obj.save()
                    addData2(A,'Member','2019-06-05',C)
                    addData2(A,'Member','2019-06-12',E)
                    addData2(A,'Member','2019-06-19',G)
                    addData2(A,'Member','2019-06-26',I)

                    addData2(A,'Member','2019-07-03',get_val(row[11]))
                    addData2(A,'Member','2019-07-10',get_val(row[13]))
                    addData2(A,'Member','2019-07-17',get_val(row[15]))
                    addData2(A,'Member','2019-07-24',get_val(row[17]))
                    addData2(A,'Member','2019-07-31',get_val(row[19]))

                    addData2(A,'Member','2019-08-07',get_val(row[22]))
                    addData2(A,'Member','2019-08-14',get_val(row[24]))
                    addData2(A,'Member','2019-08-21',get_val(row[26]))
                    addData2(A,'Member','2019-08-28',get_val(row[28]))
                    
                    addData2(A,'Member','2019-09-04',get_val(row[31]))
                    addData2(A,'Member','2019-09-11',get_val(row[33]))
                    addData2(A,'Member','2019-09-18',get_val(row[35]))
                    addData2(A,'Member','2019-09-25',get_val(row[37]))
                    
                    addData2(A,'Member','2019-10-09',get_val(row[39]))
                    addData2(A,'Member','2019-10-16',get_val(row[41]))
                    addData2(A,'Member','2019-10-23',get_val(row[43]))
                    # addData2(A,'Member','2019-08-28',get_val(row[28]))

                    
        # Data2.objects.filter(member='Member').delete()
        # m3(src,'Registration')
        def fixPoints():
            list1 = Data2.objects.all()
            for x in list1:
                if x.role == 'Speaker':
                    x.points= 3
                    x.save()

                if x.role == 'Attendance':
                    x.points= 1
                    x.save()

                if x.role in ['Ah-counter','GE','Grammarian','IE','TME','TT Evaluator','TT-master','Timer'] :
                    x.points= 2
                    x.save()

                
                    
        # fixPoints()

        def m3Guest(src,sheetname):
                wb = load_workbook(filename=src,read_only=False)
            
                ws =wb[sheetname]
                
                def get_val(cell):
                    if cell == 'None':
                        return '---'
                    if row[1].value in [None,'None']:
                        return 'xxx'
                    else:
                        if str(cell.value) == 'None':
                            return '---'
                        
                        return str(cell.value)
                rec = 0
                mem_cnt = 0
                for row in ws.rows:
                    rec += 1
                        
                    A = get_val(row[0])
                    # self.stdout.write(A)
                    B = get_val(row[1])
                    #  4 if var1 is None else var1
                    C = get_val(row[2])
                    
                    D = get_val(row[3])
                    E = get_val(row[4])
                    F = get_val(row[5])
                    G = get_val(row[6])
                    H = get_val(row[7])
                    I = get_val(row[8])
                    J = get_val(row[9])
                    K = get_val(row[10])
                
                    if B == 'Guest':
                        # mem_cnt += 1
                        # temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C+' '+ E+' '+ G
                        # self.stdout.write(temp)
                    

                        def addData2(name,member, date1, role):
                            if role:
                                obj = Data2(name=role,member=member,date1=date1,role='Attendance')
                                temp = obj.date1+ ' '+obj.member+' '+obj.name+' '+obj.role
                                if obj.name == '---':
                                    pass
                                else:
                                    self.stdout.write(temp)
                                    obj.save()
                            
                            # obj.save()
                        addData2(A,'Guest','2019-06-05',C)
                        addData2(A,'Guest','2019-06-12',E)
                        addData2(A,'Guest','2019-06-19',G)
                        addData2(A,'Guest','2019-06-26',I)

                        addData2(A,'Guest','2019-07-03',get_val(row[11]))
                        addData2(A,'Guest','2019-07-10',get_val(row[13]))
                        addData2(A,'Guest','2019-07-17',get_val(row[15]))
                        addData2(A,'Guest','2019-07-24',get_val(row[17]))
                        addData2(A,'Guest','2019-07-31',get_val(row[19]))

                        addData2(A,'Guest','2019-08-07',get_val(row[22]))
                        addData2(A,'Guest','2019-08-14',get_val(row[24]))
                        addData2(A,'Guest','2019-08-21',get_val(row[26]))
                        addData2(A,'Guest','2019-08-28',get_val(row[28]))
                        
                        addData2(A,'Guest','2019-09-04',get_val(row[31]))
                        addData2(A,'Guest','2019-09-11',get_val(row[33]))
                        addData2(A,'Guest','2019-09-18',get_val(row[35]))
                        addData2(A,'Guest','2019-09-25',get_val(row[37]))
                        
                        addData2(A,'Guest','2019-10-09',get_val(row[39]))
                        addData2(A,'Guest','2019-10-16',get_val(row[41]))
                        addData2(A,'Guest','2019-10-23',get_val(row[43]))
                        # addData2(A,'Member','2019-08-28',get_val(row[28]))
        # m3Guest(src,'Registration')
        
        def m5Member(src,sheetname):
                wb = load_workbook(filename=src,read_only=False)
                ws =wb[sheetname]
                
                def get_val(cell):
                    if cell == 'None':
                        return '---'
                    if row[1].value in [None,'None']:
                        return 'xxx'
                    else:
                        if str(cell.value) == 'None':
                            return '---'
                        
                        return str(cell.value)
                rec = 0
                mem_cnt = 0
                for row in ws.rows:
                    rec += 1
                        
                    A = get_val(row[0])
                    # self.stdout.write(A)
                    B = get_val(row[1])
                    #  4 if var1 is None else var1
                    # C = get_val(row[2])
                    
                    # D = get_val(row[3])
                    # E = get_val(row[4])
                    # F = get_val(row[5])
                    # G = get_val(row[6])
                    # H = get_val(row[7])
                    # I = get_val(row[8])
                    # J = get_val(row[9])
                    # K = get_val(row[10])
                
                    if B == 'Member':
                        # mem_cnt += 1
                        # temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C+' '+ E+' '+ G
                        # self.stdout.write(temp)
                    

                        def addData2(name,member, date1, role):
                            if role:
                                obj = Data2(name=role,member=member,date1=date1,role='Attendance')
                                temp = obj.date1+ ' '+obj.member+' '+obj.name+' '+obj.role
                                if obj.name == '---':
                                    pass
                                else:
                                    self.stdout.write(temp)
                                    # obj.save()
                            
                            # obj.save()
                            # class Person(models.Model):
                            #     club = models.ForeignKey(Club, on_delete=models.CASCADE,default=1)
                                
                            #     name = models.CharField(max_length=32) # allow same shortname
                            #     fullname = models.CharField(max_length=100) # but not fullname
                            #     is_member = models.BooleanField('Is Member ', default=False)
                            #     member_since = models.DateField('Member Since', null=True, blank=True)
                            #     note = models.CharField(max_length=100, null=True, blank=True) # but not fullname
                        club = Club.objects.get(name='Fun+')
                        p = Person(club=club,name=A,fullname=A,is_member=True)
                        p.save()
                        temp =p.club.name +" "+p.name
                        self.stdout.write(temp)
                                  
                        # addMember(A,'Guest','2019-06-05',C)
                       
        # m5Member(src,'Registration')
        dict2={2:'2019-06-05',4:'2019-06-12',6:'2019-06-19',8:'2019-06-26',11:'2019-07-03',13:'2019-07-10',15:'2019-07-17',17:'2019-07-24',19:'2019-07-31',22:'2019-08-07',24:'2019-08-14',26:'2019-08-21',28:'2019-08-28',31:'2019-09-04',33:'2019-09-11',35:'2019-09-18',37:'2019-09-25',40:'2019-10-09',42:'2019-10-16',44:'2019-10-23',46:'2019-10-30',49:'2019-11-06',51:'2019-11-13',53:'2019-11-20',}
        def m5Meeting(src,sheetname):
                wb = load_workbook(filename=src,read_only=False)
                ws =wb[sheetname]
                
                def get_val(cell):
                    if cell == 'None':
                        return '---'
                    if row[1].value in [None,'None']:
                        return 'xxx'
                    else:
                        if str(cell.value) == 'None':
                            return '---'
                        
                        return str(cell.value)
                rec = 0
                mem_cnt = 0
                for row in ws.rows:
                    rec += 1
                        
                    A = get_val(row[0])
                    # self.stdout.write(A)
                    B = get_val(row[1])
                   
                    if B == 'Member':
                     

                        club = Club.objects.get(name='Fun+')
                        p = Person.objects.get(club=club,name=A)
                        # p.save()
                        
                        for k,val in dict2.items():
                            role = get_val(row[k])
                            temp =club.name+" "+p.name+" "+str(k)+" "+val+ " "+ role
                            if role in ['Absence','---',]:
                                # temp += " ... to skip"
                                # self.stdout.write(temp)
                                pass

                            else:
                                # temp += " ... to skip"
                                if role == 'TT-master':
                                    role = 'TT Master'
                                if role == 'Ah-counter':
                                    role = 'Ah-Counter'
                                
                                
                                try:
                                    role2 = Role.objects.get(name=role)
                                except :
                                    self.stdout.write("... to fix "+ role)
                                m = Meeting(club=club,person=p,date1=val,role2=role2)
                                temp2 = m.club.name+" "+m.person.name+" "+ m.date1+" "+m.role2.name
                                m.save()
                                self.stdout.write(temp2)
                                
                                  
                        # addMember(A,'Guest','2019-06-05',C)
                       
        m5Meeting(src,'Registration')
        